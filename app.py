import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import streamlit as st

# Set up the look and feel of the web page with a sleek modern layout
st.set_page_config(
    page_title="JV Generation & Analytics Console", 
    page_icon="📊", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inject custom modern CSS to style containers, text, and metric blocks
st.markdown("""
    <style>
    /* Global Background and Global Font adjustments */
    html, body, [data-testid="stAppViewContainer"] {
        font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, Roboto, Helvetica, Arial, sans-serif;
    }
    
    /* Main Dashboard Header Custom Styling */
    .main-title {
        font-size: 2.5rem;
        font-weight: 700;
        color: #1E293B;
        margin-bottom: 0.2rem;
        letter-spacing: -0.025em;
    }
    .main-subtitle {
        font-size: 1.05rem;
        color: #64748B;
        margin-bottom: 2rem;
    }
    
    /* Sleek card look for KPIs */
    .kpi-container {
        background-color: #FFFFFF;
        padding: 1.25rem 1.5rem;
        border-radius: 12px;
        border: 1px solid #E2E8F0;
        box-shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.05), 0 1px 2px 0 rgba(0, 0, 0, 0.03);
        margin-bottom: 1rem;
    }
    .kpi-label {
        font-size: 0.85rem;
        font-weight: 600;
        color: #64748B;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 0.25rem;
    }
    .kpi-value {
        font-size: 1.75rem;
        font-weight: 700;
        color: #0F172A;
    }
    
    /* Modern Section Header Subtitles */
    .section-title {
        font-size: 1.35rem;
        font-weight: 600;
        color: #1E293B;
        margin-top: 1.5rem;
        margin-bottom: 0.5rem;
    }
    </style>
""", unsafe_allow_html=True)

# Custom Rendered Title Block
st.markdown('<div class="main-title">📊 Treasury Management Web Console</div>', unsafe_allow_html=True)
st.markdown('<div class="main-subtitle">Generate fresh templates, run automated structural audits, and export balanced journals straight to SAP.</div>', unsafe_allow_html=True)

# Initialize a session state key to handle file uploader resetting without page refreshes
if "uploader_key" not in st.session_state:
    st.session_state["uploader_key"] = 0

# ==========================================
# ENGINE 1: THE EXCEL GENERATOR
# ==========================================
def create_jv_template():
    """Builds a 300-row template in memory and returns it as a downloadable file object."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "JV Upload Form"
    ws1.views.sheetView[0].showGridLines = True

    # Styles definition
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    LABEL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    FONT_HEADER = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    FONT_LABEL = Font(name="Segoe UI", size=10, bold=True, color="1F4E78")
    FONT_BODY = Font(name="Segoe UI", size=11)
    
    THIN_BORDER = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                         top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))

    header_fields = [
        ("Document Date (YYYYMMDD):", "B1"),
        ("Posting Date (YYYYMMDD):", "B2"),
        ("Reference:", "E1"),
        ("Header Text:", "E2"),
        ("Control Total (Credits):", "H1")
    ]
    
    for label_text, cell_coord in header_fields:
        cell = ws1[cell_coord]
        cell.value = label_text
        cell.font = FONT_LABEL
        cell.fill = LABEL_FILL
        cell.alignment = Alignment(horizontal="right", vertical="center")
        
    ws1["I1"] = "=SUMIF(I7:I306, \"CR\", H7:H306)"
    ws1["I1"].number_format = "$#,##0.00"
    ws1["I1"].font = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")
    ws1["I1"].border = THIN_BORDER
    
    for input_cell in ["C1", "C2", "F1", "F2"]:
        ws1[input_cell].number_format = "@"
        ws1[input_cell].font = FONT_BODY
        ws1[input_cell].border = THIN_BORDER

    headers = ["Line No", "Fund", "GL acct", "Business area", "Functional area", "Cost Center", 
               "WBS Element", "Amount", "DR/CR", "Line Description", "Validation Status"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=6, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws1.row_dimensions[6].height = 28

    for row_idx in range(7, 307):
        ws1.row_dimensions[row_idx].height = 20
        ws1.cell(row=row_idx, column=1, value=(row_idx - 6))
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = FONT_BODY
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL
            
            if col_idx == 8:
                cell.number_format = "#######0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [1, 2, 3, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "@"
            elif col_idx in [9, 11]:
                cell.alignment = Alignment(horizontal="center")

    for col in ws1.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == "A":
            ws1.column_dimensions[col_letter].width = 10
            continue
        max_len = 0
        for cell in col:
            if cell.value is not None:
                val_str = str(cell.value)
                if val_str.startswith('='):
                    val_str = " $999,999.99 "
                max_len = max(max_len, len(val_str))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 16)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# ==========================================
# ENGINE 2: THE AUDITOR & CONVERTER
# ==========================================
def process_and_validate_jv(uploaded_file):
    """Audits the uploaded spreadsheet stream and compiles data directly into SAP text lines."""
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        if "JV Upload Form" not in wb.sheetnames:
            return False, ["Invalid Template: Missing 'JV Upload Form' sheet."], 0, 0, None

        ws = wb["JV Upload Form"]
        
        doc_date_raw = str(ws["C1"].value or "").strip()
        post_date_raw = str(ws["C2"].value or "").strip()
        reference = str(ws["F1"].value or "T1919").strip()
        header_text = str(ws["F2"].value or f"Transmittal Accrual {datetime.today().strftime('%Y')}").strip()
        
        def is_invalid_date(date_str):
            if len(date_str) != 8 or not date_str.isdigit():
                return True
            try:
                datetime.strptime(date_str, "%Y%m%d")
                return False
            except ValueError:
                return True

        if not doc_date_raw:
            return False, ["'Document Date' (Cell C1) cannot be empty!"], 0, 0, None
        if is_invalid_date(doc_date_raw):
            return False, [f"'Document Date' [{doc_date_raw}] must be exactly 8 digits (YYYYMMDD)!"], 0, 0, None

        if not post_date_raw:
            return False, ["'Posting Date' (Cell C2) cannot be empty!"], 0, 0, None
        if is_invalid_date(post_date_raw):
            return False, [f"'Posting Date' [{post_date_raw}] must be exactly 8 digits (YYYYMMDD)!"], 0, 0, None
            
        total_debit = 0.0
        total_credit = 0.0
        detailed_errors = []
        
        sap_lines = [f"H|EJ|{reference}|{doc_date_raw}|{post_date_raw}|{header_text}|11|X\n"]

        for row in range(7, 307):
            fund = str(ws.cell(row=row, column=2).value or "").strip()
            gl_account = str(ws.cell(row=row, column=3).value or "").strip()
            bus_area = str(ws.cell(row=row, column=4).value or "").strip()
            func_area = str(ws.cell(row=row, column=5).value or "").strip()
            cost_center = str(ws.cell(row=row, column=6).value or "").strip()
            wbs_element = str(ws.cell(row=row, column=7).value or "").strip()
            amount_val = ws.cell(row=row, column=8).value
            drcr_flag = str(ws.cell(row=row, column=9).value or "").strip().upper()
            description = ws.cell(row=row, column=10).value or "Transmittal Entry"

            if not fund and not gl_account and not bus_area and not func_area and not cost_center and not amount_val and drcr_flag == "":
                continue

            if fund and (len(fund) != 6 or not fund.isdigit()):
                detailed_errors.append(f"Row {row}: Fund must be exactly 6 digits (Found: '{fund}')")
                continue
                
            if not gl_account:
                detailed_errors.append(f"Row {row}: Missing GL Account")
                continue
            elif len(gl_account) != 6 or not gl_account.isdigit():
                detailed_errors.append(f"Row {row}: GL Account must be exactly 6 digits (Found: '{gl_account}')")
                continue

            if bus_area and (len(bus_area) != 4 or not bus_area.isdigit()):
                detailed_errors.append(f"Row {row}: Business Area must be exactly 4 digits (Found: '{bus_area}')")
                continue

            if func_area and (len(func_area) != 7 or not func_area.isdigit()):
                detailed_errors.append(f"Row {row}: Functional Area must be exactly 7 digits (Found: '{func_area}')")
                continue

            if cost_center and not re.match(r"^\d{8}-\d{6}$", cost_center):
                detailed_errors.append(f"Row {row}: Cost Center structure must be 8digits-6digits (Found: '{cost_center}')")
                continue

            try:
                amount = float(amount_val or 0.0)
            except ValueError:
                detailed_errors.append(f"Row {row}: Amount column has a non-numeric value.")
                continue

            if amount < 0:
                detailed_errors.append(f"Row {row}: Amount value cannot be negative.")
            elif drcr_flag not in ["DR", "CR"]:
                detailed_errors.append(f"Row {row}: Must specify 'DR' or 'CR' format criteria.")
            else:
                if drcr_flag == "DR":
                    total_debit += amount
                elif drcr_flag == "CR":
                    total_credit += amount
                
                amount_formatted = f"{amount:.2f}"
                sap_line = f"D|{row-6}|{fund}|{gl_account}|{bus_area}|{func_area}|{cost_center}|{wbs_element}|{amount_formatted}|{drcr_flag}||||||{description}\n"
                sap_lines.append(sap_line)

        if len(detailed_errors) > 0:
            return False, detailed_errors, total_debit, total_credit, None

        if abs(total_debit - total_credit) > 0.01:
            return False, ["Out of Balance! Total Debits must exactly equal Total Credits."], total_debit, total_credit, None

        txt_content = "".join(sap_lines)
        return True, [], total_debit, total_credit, txt_content

    except Exception as e:
        return False, [f"Could not parse spreadsheet data layout: {str(e)}"], 0, 0, None


# ==========================================
# PHASE 3: THE WEB USER INTERFACE (SIDEBAR & PANELS)
# ==========================================
with st.sidebar:
    st.markdown('<div style="font-size: 1.2rem; font-weight: 700; color: #1E293B; margin-bottom: 0.5rem;">🗂 Template Operations</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size: 0.9rem; color: #64748B; margin-bottom: 1rem;">Need a fresh entry form? Download a blank configured tracking worksheet.</div>', unsafe_allow_html=True)
    
    template_data = create_jv_template()
    st.download_button(
        label="📥 Download Blank Form Template",
        data=template_data,
        file_name="JV_Upload_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    st.divider()
    st.caption("System Version: 2026.WebSleek")

# Main Page Action Area
st.markdown('<div class="section-title">📁 Drag & Drop Validation Panel</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Drop your finalized entry sheet here to verify structural integrity and calculations:", 
    type=["xlsx"], 
    key=f"jv_uploader_{st.session_state['uploader_key']}",
    label_visibility="collapsed"
)

if uploaded_file is not None:
    success, errors, debits, credits, sap_txt_output = process_and_validate_jv(uploaded_file)
    
    # Modernized Custom Card Metrics Dashboard
    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([2, 2, 1.2])
    
    with col1:
        st.markdown(f"""
            <div class="kpi-container">
                <div class="kpi-label">Calculated Total Debits</div>
                <div class="kpi-value" style="color: #2563EB;">${debits:,.2f}</div>
            </div>
        """, unsafe_allow_html=True)
        
    with col2:
        st.markdown(f"""
            <div class="kpi-container">
                <div class="kpi-label">Calculated Total Credits</div>
                <div class="kpi-value" style="color: #059669;">${credits:,.2f}</div>
            </div>
        """, unsafe_allow_html=True)
        
    with col3:
        st.write("") # Spacer padding
        if st.button("🔄 Clear & Upload New JV", use_container_width=True, type="secondary"):
            st.session_state["uploader_key"] += 1
            st.rerun()
            
    st.markdown("<br>", unsafe_allow_html=True)
        
    if success:
        st.success("🎉 Perfect! Your spreadsheet passed all date constraints, digit matching structure layouts, and balances perfectly!")
        
        # Modern prominent placement button for exporting clean text data files
        st.download_button(
            label="💾 Download Ready SAP Text File (.txt)",
            data=sap_txt_output,
            file_name=f"SAP_JV_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            type="primary"
        )
    else:
        st.error("❌ Structural Validation Failed! Please resolve the following errors inside your spreadsheet and re-upload:")
        for err in errors:
            st.markdown(f"• {err}")
